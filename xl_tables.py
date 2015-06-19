import pandas as pd
import numpy as np
import openpyxl as xl

from pandas.io.excel import read_excel

html_template = """\
<!doctype html>
<html lang="en">
    <head>
    <title>{title}</title>
    <meta charset="utf-8"/>
	<meta name="viewport" content="width=device-width, user-scalable=yes, initial-scale=1.0, minimum-scale=1.0, maximum-scale=1.0" />
    </head>
    <body>
    <h2>{heading}</h2>
    <p>{table}</p>
    </body>
</html>
"""

def extract_tables(psheet, extend=True):
    psheet.index = range(len(psheet)) # This may no longer be necessary
    sheet_header = psheet.iloc[1], [1]
    # Occasional random cells have odd characters in them. I
    # haven't found a valid column with a single value in it yet ...
    for name in list(psheet.columns.values):
        if psheet[name].count() <= 1:
            del psheet[name]

    # Establish the limits of the tables
    row_counts = list(psheet.count(axis=1).values)
    endpos = len(row_counts)-1
    while endpos and row_counts[endpos] < 2:
        endpos -= 1
    startpos = 0
    while startpos < endpos-1 and (row_counts[startpos] != 1 or
                                   row_counts[startpos+1] <= 1):
        startpos += 1

    # We assume there is no table of three columns or less
    # Pandas counts the non-empty values in the row for us
    row_counts = list(psheet.count(axis=1).values)
    start_row_nums = []
    for (i, val) in enumerate(row_counts[:-1]):
        if val==1:
            if row_counts[i+1] != row_counts[i+2]:
                continue
            else:
                if row_counts[i+1] > 3:
                    start_row_nums.append(i)
    # Now we have a starting position for each table
    tables = []
    titles = []
    for start_row_num in start_row_nums:
        columns = list(psheet.iloc[start_row_num+1,:])
        titles.append(psheet.iloc[start_row_num, 0])
        end_row_num = start_row_num+1
        while row_counts[end_row_num] > 0:
            end_row_num += 1
        table = pd.DataFrame(psheet.iloc[start_row_num+2:end_row_num,:])
        table.index = range(end_row_num-start_row_num-2) # No title, no headers
        table.columns = columns
        table[columns[0]] = table[columns[0]].fillna(method="ffill")
        table.replace("-", np.NaN, inplace=True)
        if extend:
            table["Geography"] = titles[-1]
        tables.append(table)
    return sheet_header, dict(zip(titles, tables))

if __name__ == "__main__":
    import sys
    # adul-crit-care-data-eng-apr-13-mar-14-tab.xls
    pd.options.display.float_format = '{:,.0f}'.format
    #pwb = read_excel("data/gpearnextime.xlsx", sheetname=None)
    #wb = xl.load_workbook("data/gpearnextime.xlsx")
    pwb = read_excel("data/adul-crit-care-data-eng-apr-13-mar-14-tab.xlsx", sheetname=None)
    wb = xl.load_workbook("data/adul-crit-care-data-eng-apr-13-mar-14-tab.xlsx")
    tot_tbls = 0
    all_tables = {}
    for sheet_name in wb.sheetnames:
        if wb.get_sheet_by_name(sheet_name).sheet_state == "hidden":
            continue
        psheet = pwb[sheet_name]
        sheet_header, tables = extract_tables(psheet)
        print("""\
=====================================================
Sheet {}: {} tables""".format(sheet_name, len(tables)))
        for i,(tbl_title, table) in enumerate(tables.items()):
            print("{0}:  {2} x {3}: {1} ".format(i, tbl_title, *table.shape))
            tbl_html = table.to_html()
            html_out = html_template.format(heading=tbl_title, title=sheet_name, table=tbl_html)
            with open("tables/{}-{}.html".format(sheet_name.strip().replace(" ","_"), i), mode='w', encoding="utf-8") as f:
                f.write(html_out)
        tot_tbls += len(tables)
        all_tables[sheet_name] = tables
    print("""
=====================================================
    """)
    print("\n\n\nFOR A TOTAL OF {} EXTRACTED TABLES".format(tot_tbls))

print(end="") # just here for debugging